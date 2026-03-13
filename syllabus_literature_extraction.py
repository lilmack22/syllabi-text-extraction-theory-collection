"""
syllabus_literature_extraction.py

Reads every PDF in "Syllabi to Draw From", calls GROBID to extract
bibliographic references, deduplicates readings that appear in multiple
syllabi, and writes a structured Excel spreadsheet called
literature_from_selected_syllabi.xlsx.

Required packages:
    pip install requests pandas openpyxl lxml

GROBID must be running at http://localhost:8070 before this script is run.
Course metadata (title, professor, term, year) is read from
syllabi_metadata.xlsx if it exists (produced by syllabi_metadata_extraction.py).
"""

import re
import logging
import requests
import pandas as pd
from datetime import date
from pathlib import Path
from lxml import etree

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

GROBID_BASE_URL = "http://localhost:8070"
SYLLABI_FOLDER  = Path(__file__).parent / "Syllabi to Draw From"
METADATA_XLSX   = Path(__file__).parent / "syllabi_metadata.xlsx"
OUTPUT_XLSX     = Path(__file__).parent / "literature_from_selected_syllabi.xlsx"

TEI_NS = "http://www.tei-c.org/ns/1.0"
NS     = {"tei": TEI_NS}

# Token-level Jaccard threshold for deciding two titles are the same reading.
# Increase toward 1.0 for stricter (fewer merges), decrease for looser.
TITLE_SIM_THRESHOLD = 0.82

COLUMNS = [
    "Original Citation",
    "Author/s",
    "Year",
    "Title",
    "Journal/Publication",
    "Issue/Pages",
    "Link to Paper",
    "Source Type",
    "Date Added",
    "Class/es where listed on syllabus",
]

logging.basicConfig(level=logging.INFO, format="%(levelname)s | %(message)s")
log = logging.getLogger(__name__)

TODAY = date.today().isoformat()


# ---------------------------------------------------------------------------
# GROBID helpers
# ---------------------------------------------------------------------------

def check_grobid_alive() -> bool:
    try:
        r = requests.get(f"{GROBID_BASE_URL}/api/isalive", timeout=10)
        return r.status_code == 200
    except requests.RequestException:
        return False


def call_grobid_fulltext(pdf_path: Path) -> str | None:
    """
    POST to processFulltextDocument with citation consolidation.
    Returns raw TEI/XML string or None on failure.
    """
    url = f"{GROBID_BASE_URL}/api/processFulltextDocument"
    try:
        with open(pdf_path, "rb") as fh:
            resp = requests.post(
                url,
                files={"input": (pdf_path.name, fh, "application/pdf")},
                data={
                    "consolidateCitations": "1",   # enrich with DOI / URLs
                    "includeRawCitations":  "1",   # preserve original citation text
                },
                timeout=180,
            )
        if resp.status_code == 200:
            return resp.text
        log.warning("  GROBID fulltext HTTP %s for %s", resp.status_code, pdf_path.name)
    except requests.RequestException as exc:
        log.error("  GROBID fulltext request failed for %s: %s", pdf_path.name, exc)
    return None


# ---------------------------------------------------------------------------
# TEI / XML helpers
# ---------------------------------------------------------------------------

def parse_tei(xml_text: str, label: str = "") -> etree._Element | None:
    """Parse TEI string; falls back to lxml recovery mode on syntax errors."""
    if not xml_text or not xml_text.strip():
        return None
    try:
        return etree.fromstring(xml_text.encode("utf-8"))
    except etree.XMLSyntaxError:
        pass
    try:
        parser = etree.XMLParser(recover=True)
        return etree.fromstring(xml_text.encode("utf-8"), parser)
    except Exception as exc:
        log.warning("  TEI parse failed (%s): %s", label, exc)
    return None


def _text(node) -> str:
    """Join all descendant text nodes, stripped."""
    if node is None:
        return ""
    return " ".join(node.itertext()).strip()


# ---------------------------------------------------------------------------
# Reference extraction
# ---------------------------------------------------------------------------

def extract_references(root: etree._Element) -> list[dict]:
    """
    Find all <biblStruct> elements inside <listBibl> (the bibliography /
    reading-list section that GROBID generates) and parse each one.
    """
    refs = []
    for bibl in root.xpath(".//tei:listBibl//tei:biblStruct", namespaces=NS):
        ref = _parse_biblstruct(bibl)
        if ref.get("title"):
            refs.append(ref)
    return refs


def _parse_biblstruct(bibl: etree._Element) -> dict:
    """Return a dict of bibliographic fields for one <biblStruct>."""

    # Raw citation text stored by GROBID when includeRawCitations=1
    raw_note = bibl.find(f".//{{{TEI_NS}}}note[@type='raw_reference']")
    raw_citation = _text(raw_note) if raw_note is not None else ""

    analytic = bibl.find(f"{{{TEI_NS}}}analytic")
    monogr   = bibl.find(f"{{{TEI_NS}}}monogr")

    # Title: article title from <analytic>, book/journal title from <monogr>
    title = ""
    if analytic is not None:
        for t in analytic.xpath("tei:title", namespaces=NS):
            title = _text(t)
            if title:
                break
    if not title and monogr is not None:
        for t in monogr.xpath("tei:title", namespaces=NS):
            title = _text(t)
            if title:
                break

    # Authors: prefer <analytic> authors (article authors),
    # fall back to <monogr> authors (book authors)
    authors = _extract_authors(analytic) or _extract_authors(monogr)

    # Year
    year = ""
    for date_node in bibl.xpath(
        ".//tei:imprint/tei:date | .//tei:date", namespaces=NS
    ):
        candidate = date_node.get("when", "") or _text(date_node)
        m = re.search(r"\b(1[89]\d\d|20[0-3]\d)\b", candidate)
        if m:
            year = m.group()
            break

    # Journal / publication name
    journal = ""
    if monogr is not None:
        # Prefer titles explicitly typed as journal (j), monograph (m), or series (s)
        for t in monogr.xpath("tei:title", namespaces=NS):
            lvl = t.get("level", "")
            if lvl in ("j", "m", "s"):
                journal = _text(t)
                if journal:
                    break
        if not journal:
            for t in monogr.xpath("tei:title", namespaces=NS):
                candidate = _text(t)
                if candidate and candidate != title:
                    journal = candidate
                    break

    # Volume / issue / pages
    issue_pages = _extract_issue_pages(bibl)

    # DOI → URL; fallback to other identifier types
    link = ""
    for idno in bibl.xpath(".//tei:idno", namespaces=NS):
        id_type = (idno.get("type") or "").lower()
        val = _text(idno)
        if id_type == "doi" and val:
            link = f"https://doi.org/{val}"
            break
        if id_type in ("url", "arxiv", "pmid") and val and not link:
            link = val

    source_type = _infer_source_type(bibl, journal, analytic)

    return {
        "original_citation": raw_citation,
        "authors":           authors,
        "year":              year,
        "title":             title,
        "journal":           journal,
        "issue_pages":       issue_pages,
        "link":              link,
        "source_type":       source_type,
    }


def _extract_authors(container: etree._Element | None) -> str:
    """Extract all author names from a TEI container element."""
    if container is None:
        return ""
    names = []
    for author in container.xpath(".//tei:author", namespaces=NS):
        forenames = " ".join(
            n.text.strip()
            for n in author.xpath(".//tei:forename", namespaces=NS) if n.text
        )
        surname = " ".join(
            n.text.strip()
            for n in author.xpath(".//tei:surname", namespaces=NS) if n.text
        )
        full = f"{forenames} {surname}".strip()
        if not full:
            pn = author.find(f"{{{TEI_NS}}}persName")
            full = _text(pn) if pn is not None else _text(author)
        if full:
            names.append(full)
    return "; ".join(names)


def _extract_issue_pages(bibl: etree._Element) -> str:
    """Build a Vol./No./pp. string from <biblScope> elements."""
    parts = []
    vol = bibl.xpath(".//tei:imprint/tei:biblScope[@unit='volume']", namespaces=NS)
    iss = bibl.xpath(".//tei:imprint/tei:biblScope[@unit='issue']",  namespaces=NS)
    pps = bibl.xpath(".//tei:imprint/tei:biblScope[@unit='page']",   namespaces=NS)
    if vol:
        parts.append(f"Vol. {_text(vol[0])}")
    if iss:
        parts.append(f"No. {_text(iss[0])}")
    if pps:
        page_text = _text(pps[0])
        if not page_text:
            fr = pps[0].get("from", "")
            to = pps[0].get("to",   "")
            if fr and to:
                page_text = f"{fr}–{to}"
            elif fr:
                page_text = fr
        if page_text:
            parts.append(f"pp. {page_text}")
    return ", ".join(parts)


def _infer_source_type(
    bibl: etree._Element,
    journal: str,
    analytic: etree._Element | None,
) -> str:
    """
    Heuristic source-type classification:
      - <analytic> present + journal name  →  Journal Article
      - <monogr> with publisher            →  Book
      - Otherwise                          →  White Paper
    """
    if analytic is not None:
        analytic_titles = analytic.xpath("tei:title", namespaces=NS)
        if analytic_titles and journal:
            return "Journal Article"
    monogr = bibl.find(f"{{{TEI_NS}}}monogr")
    if monogr is not None:
        imprint = monogr.find(f"{{{TEI_NS}}}imprint")
        if imprint is not None:
            publisher = imprint.find(f"{{{TEI_NS}}}publisher")
            if publisher is not None and _text(publisher):
                return "Book"
    if journal:
        return "Journal Article"
    return "White Paper"


# ---------------------------------------------------------------------------
# Deduplication
# ---------------------------------------------------------------------------

def _normalise(title: str) -> str:
    """Lowercase, strip punctuation/articles, collapse whitespace."""
    t = title.lower()
    t = re.sub(r"[^\w\s]", " ", t)
    t = re.sub(r"\b(a|an|the)\b", " ", t)
    return re.sub(r"\s+", " ", t).strip()


def _jaccard(a: str, b: str) -> float:
    """Token-level Jaccard similarity."""
    sa = set(a.split())
    sb = set(b.split())
    if not sa or not sb:
        return 0.0
    return len(sa & sb) / len(sa | sb)


def merge_references(all_refs: list[tuple[str, dict]]) -> list[dict]:
    """
    Deduplicate references across syllabi.

    all_refs: list of (class_label, ref_dict) tuples.
    Returns a list of unique ref dicts.  Each dict has a 'classes' key
    containing a list of all class labels that assigned that reading.
    """
    merged:     list[dict] = []
    norm_index: list[str]  = []   # normalised title for each merged entry

    for class_label, ref in all_refs:
        norm = _normalise(ref["title"])

        best_idx   = -1
        best_score = 0.0
        for i, existing_norm in enumerate(norm_index):
            score = _jaccard(norm, existing_norm)
            if score > best_score:
                best_score = score
                best_idx   = i

        if best_score >= TITLE_SIM_THRESHOLD and best_idx >= 0:
            # Duplicate: merge class label into existing entry
            classes = merged[best_idx]["classes"]
            if class_label and class_label not in classes:
                classes.append(class_label)
        else:
            # New unique entry
            entry = dict(ref)
            entry["classes"]    = [class_label] if class_label else []
            entry["date_added"] = TODAY
            merged.append(entry)
            norm_index.append(norm)

    return merged


# ---------------------------------------------------------------------------
# Syllabus metadata loader
# ---------------------------------------------------------------------------

def load_class_labels(metadata_path: Path) -> dict[str, str]:
    """
    Read syllabi_metadata.xlsx and return a dict:
        PDF filename → class label string

    Label format: '"Course Title" Professor Name/s – Term Year'
    (e.g. '"The Logics of Violence" Beatriz Magaloni – Fall 2025')

    If the metadata file does not exist, returns an empty dict and the
    raw filename will be used as the class label.
    """
    if not metadata_path.exists():
        log.warning(
            "Metadata file not found at %s. "
            "Run syllabi_metadata_extraction.py first for richer class labels. "
            "Falling back to raw PDF filenames.",
            metadata_path,
        )
        return {}

    df = pd.read_excel(metadata_path)

    col_pdf   = "Original name of syllabus PDF"
    col_title = "Course Title"
    col_prof  = "Course Professors"
    col_term  = "Term (Spring, Winter, etc) the Course was Taught"
    col_year  = "Year the Course was taught"

    labels: dict[str, str] = {}
    for _, row in df.iterrows():
        pdf_name = str(row.get(col_pdf, "")).strip()
        if not pdf_name:
            continue

        title = str(row.get(col_title, "")).strip()
        prof  = str(row.get(col_prof,  "")).strip()
        term  = str(row.get(col_term,  "")).strip()
        year  = str(row.get(col_year,  "")).strip()

        # '"Title" Professors – Term Year'
        name_part = f'"{title}"' if title else ""
        if prof:
            name_part = f"{name_part} {prof}".strip()

        term_year = " ".join(filter(None, [term, year]))
        if term_year:
            label = f"{name_part} – {term_year}" if name_part else term_year
        else:
            label = name_part or pdf_name

        labels[pdf_name] = label

    return labels


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")   # dark navy
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
BODY_FONT   = Font(name="Calibri", size=10)
ALT_FILL    = PatternFill("solid", fgColor="D6E4F0")   # light blue
WRAP        = Alignment(wrap_text=True, vertical="top")
THIN_SIDE   = Side(style="thin", color="B0C4DE")
CELL_BORDER = Border(
    left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE
)

# Approximate column widths (characters)
COL_WIDTHS = {
    "A": 45,   # Original Citation
    "B": 30,   # Author/s
    "C": 8,    # Year
    "D": 55,   # Title
    "E": 32,   # Journal/Publication
    "F": 22,   # Issue/Pages
    "G": 38,   # Link to Paper
    "H": 18,   # Source Type
    "I": 14,   # Date Added
    "J": 65,   # Class/es
}


def write_excel(
    merged: list[dict],
    all_class_labels: list[str],
    out_path: Path,
) -> None:
    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------ #
    #  Sheet 1: Literature                                                 #
    # ------------------------------------------------------------------ #
    ws = wb.active
    ws.title = "Literature"

    # Header row
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = WRAP
        cell.border    = CELL_BORDER
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # Data rows
    for row_idx, ref in enumerate(merged, start=2):
        classes_str = "; ".join(ref.get("classes", []))
        values = [
            ref.get("original_citation", ""),
            ref.get("authors",           ""),
            ref.get("year",              ""),
            ref.get("title",             ""),
            ref.get("journal",           ""),
            ref.get("issue_pages",       ""),
            ref.get("link",              ""),
            ref.get("source_type",       ""),
            ref.get("date_added",        TODAY),
            classes_str,
        ]
        row_fill = ALT_FILL if row_idx % 2 == 0 else None
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = BODY_FONT
            cell.alignment = WRAP
            cell.border    = CELL_BORDER
            if row_fill:
                cell.fill = row_fill
        ws.row_dimensions[row_idx].height = 45

    # ------------------------------------------------------------------ #
    #  Sheet 2: _ClassList  (reference sheet used by data validation)     #
    # ------------------------------------------------------------------ #
    ws_cls = wb.create_sheet("_ClassList")
    ws_cls.cell(row=1, column=1, value="Available Classes").font = Font(
        bold=True, name="Calibri", size=11
    )
    sorted_classes = sorted(all_class_labels)
    for i, label in enumerate(sorted_classes, start=2):
        ws_cls.cell(row=i, column=1, value=label).font = BODY_FONT
    ws_cls.column_dimensions["A"].width = 70

    # ------------------------------------------------------------------ #
    #  Data Validation: Source Type (column H)                            #
    # ------------------------------------------------------------------ #
    last_row = max(len(merged) + 1, 2)
    extra    = 500   # leave room for rows added later

    dv_type = DataValidation(
        type="list",
        formula1='"Book,Journal Article,White Paper,Working Paper,Report,Book Chapter"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_type.sqref            = f"H2:H{last_row + extra}"
    dv_type.prompt           = "Choose source type"
    dv_type.promptTitle      = "Source Type"
    dv_type.showErrorMessage = False   # allow free-text overrides
    ws.add_data_validation(dv_type)

    # ------------------------------------------------------------------ #
    #  Data Validation: Classes column (column J)                         #
    #                                                                      #
    #  Excel does not support native multi-select dropdowns.  We attach   #
    #  a list validation pointing to _ClassList so a single class can be  #
    #  selected from the dropdown.  For readings assigned to multiple      #
    #  classes the cell contains a "; "-separated string (pre-filled by   #
    #  the script); showErrorMessage=False means Excel will not reject     #
    #  those multi-value strings.                                          #
    # ------------------------------------------------------------------ #
    if all_class_labels:
        class_formula = f"_ClassList!$A$2:$A${len(sorted_classes) + 1}"
        dv_class = DataValidation(
            type="list",
            formula1=class_formula,
            allow_blank=True,
            showDropDown=False,
        )
        dv_class.sqref            = f"J2:J{last_row + extra}"
        dv_class.prompt           = (
            "Select a class from the dropdown, or type multiple classes "
            "separated by '; ' for readings that appear in more than one course."
        )
        dv_class.promptTitle      = "Class/es"
        dv_class.showErrorMessage = False   # allow multi-value strings
        ws.add_data_validation(dv_class)

    wb.save(out_path)
    log.info("Saved spreadsheet → %s", out_path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    if not check_grobid_alive():
        log.error(
            "GROBID server is not responding at %s. "
            "Please start GROBID before running this script.",
            GROBID_BASE_URL,
        )
        return

    pdf_files = sorted(SYLLABI_FOLDER.glob("*.pdf"))
    if not pdf_files:
        log.error("No PDF files found in: %s", SYLLABI_FOLDER)
        return

    log.info("Found %d PDF(s) in '%s'", len(pdf_files), SYLLABI_FOLDER)

    # Load class labels from the previously generated metadata sheet
    class_labels = load_class_labels(METADATA_XLSX)
    all_class_label_set: set[str] = set()

    # Collect (class_label, ref_dict) pairs across all syllabi
    all_refs: list[tuple[str, dict]] = []

    for pdf_path in pdf_files:
        filename    = pdf_path.name
        class_label = class_labels.get(filename, filename)
        all_class_label_set.add(class_label)

        log.info("Processing: %s  →  class label: %s", filename, class_label)

        xml_text = call_grobid_fulltext(pdf_path)
        if not xml_text:
            log.warning("  No GROBID response; skipping %s", filename)
            continue

        root = parse_tei(xml_text, label=filename)
        if root is None:
            log.warning("  Could not parse TEI for %s; skipping", filename)
            continue

        refs = extract_references(root)
        log.info("  Extracted %d reference(s)", len(refs))
        for ref in refs:
            all_refs.append((class_label, ref))

    log.info("Total raw references across all syllabi: %d", len(all_refs))

    merged = merge_references(all_refs)
    log.info("After deduplication: %d unique reference(s)", len(merged))

    write_excel(merged, list(all_class_label_set), OUTPUT_XLSX)

    print("\n" + "=" * 58)
    print("  LITERATURE EXTRACTION SUMMARY")
    print("=" * 58)
    print(f"  PDFs processed             : {len(pdf_files)}")
    print(f"  Raw references extracted   : {len(all_refs)}")
    print(f"  Unique references (deduped): {len(merged)}")
    print(f"  Output spreadsheet         : {OUTPUT_XLSX.name}")
    print("=" * 58 + "\n")


if __name__ == "__main__":
    main()
